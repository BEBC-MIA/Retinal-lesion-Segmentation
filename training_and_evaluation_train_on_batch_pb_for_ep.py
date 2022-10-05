#!/usr/bin/env python
# coding: utf-8

# In[2]:

'''
For Multi-classes Training(DDR,IDRiD,classes=5)

change iou

'''
#from unet import unet, gn_unet,ps_gn_unet
#from pr_resunet import *
from models.unet_plusplus import *
from models.keras_data import *
from models.model import Deeplabv3
from utils_for_2classes import *
from models.loss_function import *
from evaluate import *
import math
import xlwt
from sklearn.metrics import f1_score
from tensorflow.python.keras.callbacks import TensorBoard
from tensorflow_addons.optimizers.weight_decay_optimizers import AdamW
from models.GroupNormalization import GroupNormalization
from pixle_shuffling import*
import cv2
import os
import tensorflow.keras.backend as K
from models.Unet3_Plus import unet3_plus
import openpyxl
import time
'''
gpu = "0,1"
os.environ["CUDA_VISIBLE_DEVICES"] = gpu
config = tf.compat.v1.ConfigProto()
config.gpu_options.allow_growth = True
gpu_num = len(gpu.split(','))
strategy = tf.distribute.MirroredStrategy()
'''
'''
def save_to_exel(x_list,y_list,write_path):
    xls = xlwt.Workbook()
    sht1 = xls.add_sheet("Sheet1")
    sht1.write(0, 0, 'recall')
    sht1.write(0, 1, 'precision')
    for i,x in enumerate(x_list):       
        sht1.write(i+1, 0, x)
        sht1.write(i+1, 1, y_list[i])
    xls.save(write_path)
'''
def save_to_exel(x_list,y_list,write_path):
    #xls = xlwt.Workbook()
    print(len(x_list))
    xls = openpyxl.Workbook()
    sht1 = xls.create_sheet("Sheet1")
    sht1.cell(1, 1, 'recall')
    sht1.cell(1, 2, 'precision')
    for i,x in enumerate(x_list):       
        sht1.cell(i+2, 1, x)
        sht1.cell(i+2, 2, y_list[i])
    xls.save(write_path)

def data_generator(classes = 5,
                    use_pb=True,
                   folder_dict=None,
                   target_size = (800,800),
                   train_batch_size = 2,
                   val_batch_size = 2,
                   flag_multi_class = True,
                   train_path = None,
                   val_path = None,
                   save_path = None,
                   extra_aug = True):
    """
        choosing datasets

    """

    if not os.path.exists(save_path):  # 创建文件夹
        os.makedirs(save_path)
    if flag_multi_class:
        classes = classes
        if classes == 1:
            print('classes must > 1')
    else:
        classes = 1

    """
    Generator config

    """
    train_args = dict(rotation_range=0,
                      width_shift_range=0.015,
                      height_shift_range=0.015,
                      zoom_range=0,
                      vertical_flip=True,
                      horizontal_flip=True,
                      channel_shift_range=30,
                      fill_mode='constant')

    train_label_args = dict(rotation_range=0,
                            width_shift_range=0.015,
                            height_shift_range=0.015,
                            zoom_range=0,
                            vertical_flip=True,
                            horizontal_flip=True,
                            fill_mode='constant')

    val_args = dict(rotation_range=0,
                    width_shift_range=0,
                    height_shift_range=0,
                    zoom_range=0,
                    horizontal_flip=False,
                    fill_mode='constant')

    val_label_args = dict(rotation_range=0,
                          width_shift_range=0,
                          height_shift_range=0,
                          zoom_range=0,
                          horizontal_flip=False,
                          fill_mode='constant')
    extra_aug_args = dict(zoom_range=None, rotation_range=90, elastic_trans=None)
    if use_pb:
        img_folder = folder_dict[0][0]
        label_folder = folder_dict[0][1]
    else:
        img_folder = folder_dict[1][0]
        label_folder = folder_dict[1][1]
    trainGene = trainGenerator(train_batch_size,
                               train_args,
                               train_label_args,
                               train_path,
                               img_folder,
                               label_folder,
                               vs_label_folder='vessel_mask_zoom_hd',
                               od_label_folder='od_mask_zoom_hd',
                               shuffle=True,
                               flag_multi_class=flag_multi_class,
                               num_class=classes,
                               save_to_dir=None,
                               target_size=target_size,
                               extra_aug=extra_aug,
                               extra_aug_dict=extra_aug_args
                               )

    valGene = trainGenerator(val_batch_size,
                             val_args,
                             val_label_args,
                             val_path,
                             'image_zoom_hd',
                             'label_zoom_hd',
                             shuffle=False,
                             flag_multi_class=flag_multi_class,
                             num_class=classes,
                             save_to_dir=None,
                             target_size=target_size)
    return trainGene, valGene

def build_model(classes = 5,
                target_size = (420,420),
                img_channel = 3,
                epochs = None,
                learning_rate = 0.0001,
                weight_decay_rate = 0.00005,
                model_name = None,
                supervision = True,
                use_ps = True,
                normalize = 'gn',
                loss_f = None):

    w, h = target_size
    if classes != 1:  # choosing activation
        activation = 'softmax'
    else:
        activation = 'sigmoid'
    # model = Deeplabv3(input_shape=(w,h, 3), classes=classes,OS=16)
    # model= tinyresunet(input_size=(w,h,3),classes=classes,activation='softmax',skip_type='add')
    # model= pr_resunet(input_size=(w,h,3),classes=classes,
    #                 activation='softmax',supervision=False,skip_type='add')
    #with tf.device('/cpu:0'):# 使用多GPU时，先在CPU上初始化模型
    if model_name == 'unet_plusplus':
    
        model = res_unet_plusplus(input_size = (w,h,img_channel),
                                  classes = classes,
                                  activation = activation,
                                  supervision = supervision,
                                  weight_fusion = False,
                                  use_fpn = False,
                                  use_ps = use_ps,
                                  ex_supervision = False,
                                  normalize = normalize)
    elif model_name == 'original_unet_plusplus':
        model = original_unet_plusplus(input_size = (w,h,img_channel),
                                      classes = classes,
                                      activation = activation,
                                      supervision = False,
                                      weight_fusion = False,                           
                                      use_fpn = False,
                                      use_ps = False,
                                      ex_supervision = False,
                                      normalize = normalize)
    elif model_name == 'deeplabv3':
        model = Deeplabv3(input_shape = (w, h, img_channel),
                              classes = classes,
                              OS = 8)
    elif model_name == 'unet':
        model = unet(input_size = (w,h,img_channel),
                         classes = classes,
                         activation = activation,
                         use_fpn=False,
                         use_ps = False,
                         normalize = normalize)
    elif model_name == 'unet3_plus':
        model = unet3_plus((w,h,img_channel), n_class=classes)
    # model= original_unet_plusplus(input_size=(w,h,3),classes=classes,activation='softmax',
    #                    supervision=False,weight_fusion = False,use_fpn=False,use_ps=False,ex_supervision=False,normalize = 'gn')
    # model= resunet_plusplus(input_size=(w,h,3),classes=classes,activation='softmax',
    #                     supervision=True,weight_fusion = False,use_fpn=False,use_ps=True,ex_supervision=False,normalize = 'gn')

    # model= gn_unet(input_size=(w,h,3),classes=classes,activation='softmax',supervision=False)

    # model= ps_gn_unet(input_size=(w,h,3),classes=classes,activation='softmax',supervision=False,ps=True)
    # model= unet(input_size=(w,h,3),classes=classes,activation='softmax',supervision=False)

    if epochs != None:
        epoch_list = [epochs - int(epochs * 0.2)]
    else:
        epoch_list = [12, 32]
    print(epoch_list)
    step_decay_lr = step_decay(epoch_list=epoch_list,  # [36,45]
                               decay_factor=0.5)
    reduce_lr = ReduceLROnPlateau(monitor='val_loss',
                                  patience=8,
                                  factor=0.2,
                                  mode='min',
                                  min_lr=1e-6,
                                  cooldown=1)

    """
    Optimizer config：

    """
    if learning_rate != None:
        lr = learning_rate
    else:
        lr = 0.0001
    if weight_decay_rate != None:
        wdr = weight_decay_rate
    else:
        wdr = 0.0001
    #opt = Adam(learning_rate=lr)
    #opt = RAdam(lr=lr)
    opt = AdamW(learning_rate=lr, weight_decay=wdr)
    #opt = tf.keras.optimizers.SGD(learning_rate=lr,momentum=0.9)
    #opt = tf.keras.optimizers.Nadam(lr=lr)
    #opt = SGD(lr=lr,momentum=0.9)
    """
    Loss functions:

    """

    if loss_f == 'dice_CE':
        loss = tversky_CEv2_loss_fun(alpha = [0.5, 0.5, 0.5],
                                     beta = [0.5, 0.5, 0.5],
                                     w = [0.1, 0.9, 0.9],
                                     factor = 0.4
                                     )
                                    

    elif loss_f == 'CE':
        loss = 'categorical_crossentropy'
    return model, opt, loss

def evl(results,
        method='PR',
        threshold_num=33,
        classes=5,
        target_size=(420, 420),
        task=None,
        flag_multi_class=True,
        groundtruth_path=None,
        save_path=None):
    if flag_multi_class:
        classes = classes
        if classes == 1:
            print('classes must > 1')
    else:
        classes = 1
    if classes == 3 and task == 'ex_ma':
        classes_list = ['EX', 'MA']
    y_true = get_y_true(groundtruth_path, classes, target_size=target_size,
                        task=task)  # transforming every image into array and saving in a list
    y_pred = get_y_pred(results, classes, target_size=target_size)
    au_dict = {}

    """
    evaluation methods

    """
    if classes == 1:

        # 二分类，sigmoid
        if method == 'ROC':
            precision, recall, auc = ROC(y_pred[0], y_true[0], classes_list[0])
            plt.savefig(os.path.join(save_path, classes_list[0] + '_ROC.png'))
            au_dict[classes_list[0] + '_AUC'] = auc
        elif method == 'PR':
            print(np.where(y_pred[0] > 0.5)[0])
            print(np.where(y_true[0] != 0)[0])
            precision, recall, aupr = AUPR(y_pred[0], y_true[0], classes_list[0])
            plt.savefig(os.path.join(save_path, classes_list[0] + '_PR.png'))
            au_dict[classes_list[0] + '_AUPR'] = aupr

        elif method == 'all':
            precision, recall, aupr = AUPR(y_pred[0], y_true[0], classes_list[0], threshold_num=threshold_num)
            # precision,recall,aupr = AUPR_auto(y_pred[i],y_true[i],classname)
            plt.savefig(os.path.join(save_path, classes_list[0] + '_PR.png'))
            au_dict[classes_list[0] + '_AUPR'] = aupr
            save_to_exel(recall, precision, os.path.join(save_path, classes_list[0] + '_PR.xls'))
            tp, fp, auc = ROC_auto(y_pred[0], y_true[0], classes_list[0])
            plt.savefig(os.path.join(save_path, classes_list[0] + '_ROC.png'))
            au_dict[classes_list[0] + '_AUC'] = auc
            # save_to_exel(fp,tp,os.path.join(save_path,'ROC.xls'))
            print('evaluation:', au_dict)
            compute_pr_f1(y_pred, y_true, class_num=classes, class_list=classes_list)  # compute pr,recall,f1
            compute_dice(save_path, groundtruth_path, classes=classes,task=task)  # compute average dice of all val images



    else:
        # 多分类，softmax
        sum_AUPR = 0
        sum_AUC = 0
        if method == 'ROC':  # Area under ROC curve (AUC)
            for i, classname in enumerate(classes_list):
                tp, fp, auc = ROC_auto(y_pred[i + 1], y_true[i], classname)
                sum_AUC = sum_AUC + auc
                save_to_exel(fp, tp, os.path.join(save_path, classname + '_ROC.xls'))
                au_dict[classname + '_AUC'] = auc
            mAUC = sum_AUC / (classes - 1)
            au_dict['mAUC'] = mAUC
            print('evaluation:', au_dict)
        elif method == 'PR':  # Area under PR curve (AUPR)
            for i, classname in enumerate(classes_list):
                #precision, recall, aupr = AUPR(y_pred[i + 1], y_true[i], classname, threshold_num=threshold_num)
                precision, recall, aupr = AUPR_auto(y_pred[i + 1], y_true[i], classname)
                plt.savefig(os.path.join(save_path, classname + '_PR.png'))
                sum_AUPR = sum_AUPR + aupr
                au_dict[classname + '_AUPR'] = aupr
            mAUPR = sum_AUPR / (classes - 1)     #-background
            au_dict['mAUPR'] = mAUPR
            print('evaluation:', au_dict)
        elif method == 'f1':  # F1-score
            print("f1-score: ")
            f1_dict = compute_pr_f1(y_pred, y_true, class_num=classes)  # compute pr,recall,f1
            f = open(os.path.join(save_path,'f1.txt'), 'w', encoding='utf-8')  # 以'w'方式打开文件
            for k, v in f1_dict.items():  # 遍历字典中的键值
                s2 = str(v)  # 把字典的值转换成字符型
                f.write(k + '\n')  # 键和值分行放，键在单数行，值在双数行
                f.write(s2 + '\n')
            f.close()  # 关闭文件
        elif method == 'dice':  # Dice-score
            print("dice-score: ")
            compute_dice(save_path, groundtruth_path, classes=classes,
                         task=task)  # compute average dice of all val images
        elif method == 'all':  # AUC, AUPR, F1-score, Dice-score
            for i, classname in enumerate(classes_list):
                #precision, recall, aupr = AUPR(y_pred[i + 1], y_true[i], classname, threshold_num=threshold_num)
                precision,recall,aupr = AUPR_auto(y_pred[i+1],y_true[i],classname)
                precision = precision[::-1]
                recall = recall[::-1]
                precision = precision[0:precision.size:100].tolist()
                recall = recall[0:recall.size:100].tolist()
                precision = precision[::-1]
                recall = recall[::-1]
                #np.save(os.path.join(save_path, classname + '_precision.npy'),precision)
                #np.save(os.path.join(save_path, classname + '_recall.npy'),recall)
                plt.savefig(os.path.join(save_path, classname + '_PR.png'))
                au_dict[classname + '_AUPR'] = aupr
                sum_AUPR = sum_AUPR + aupr
                save_to_exel(recall, precision, os.path.join(save_path, classname + '_PR.xlsx'))
                tp, fp, auc = ROC_auto(y_pred[i + 1], y_true[i], classname)
                tp = np.array(tp)
                fp = np.array(fp)
                np.save(os.path.join(save_path, classname + '_tp.npy'),tp)
                np.save(os.path.join(save_path, classname + '_fp.npy'),fp)
                sum_AUC = sum_AUC + auc
                #save_to_exel(fp, tp, os.path.join(save_path, classname + '_ROC.xls'))
                plt.savefig(os.path.join(save_path, classname + '_ROC.png'))
                au_dict[classname + '_AUC'] = auc
                # save_to_exel(fp,tp,os.path.join(save_path,'ROC.xls'))
            mAUPR = sum_AUPR / (classes - 1)
            au_dict['mAUPR'] = mAUPR
            mAUC = sum_AUC / (classes - 1)
            au_dict['mAUC'] = mAUC
            print('evaluation:', au_dict)
            f1_dict = compute_pr_f1(y_pred, y_true, class_num=classes)  # compute pr,recall,f1
            f = open(os.path.join(save_path,'f1.txt'), 'w', encoding='utf-8')  # 以'w'方式打开文件
            for k, v in f1_dict.items():  # 遍历字典中的键值
                s2 = str(v)  # 把字典的值转换成字符型
                f.write(k + '\n')  # 键和值分行放，键在单数行，值在双数行
                f.write(s2 + '\n')
            f.close()  # 关闭文件
            f = open(os.path.join(save_path,'auc.txt'), 'w', encoding='utf-8')  # 以'w'方式打开文件
            for k, v in au_dict.items():  # 遍历字典中的键值
                s2 = str(v)  # 把字典的值转换成字符型
                f.write(k + '\n')  # 键和值分行放，键在单数行，值在双数行
                f.write(s2 + '\n')
            f.close()  # 关闭文件

            
        else:
            print("error")
    return au_dict

def train(
          model, opt, loss,
          trainGene,
          valGene,
          GPU_id = "0",
          step = 1,
          val_metrics = True,
          train_strategy = 'step_decay',
          classes = 5,
          target_size = (420,420),
          train_batch_size = 1,
          val_batch_size = 2,
          val_num = None,
          train_num = None,
          epochs = None,
          bg_epoch  = 0,
          learning_rate = 0.0001,
          weight_decay_rate = 0.0001,
          save_result = True,
          task = None,
          flag_multi_class = True,
          log_name = None,
          log_path = None,
          test_path = None,
          groundtruth_path = None,
          save_path = None,
          load_dir_initial = None):
    '''
    def write_log(callback, logs_dict, epoch):
        for key, value in logs_dict.items():
            summary = tf.compat.v1.Summary()
            summary_value = summary.value.add()
            summary_value.simple_value = value
            summary_value.tag = key
            #callback.on_epoch_end(epoch,logs={key:value})
            callback.writer.add_summary(summary, epoch)
            callback.writer.flush()

    '''
    load_dir = save_path + '/ep_lesion_final.hdf5'
    summary_writer = tf.summary.create_file_writer(os.path.join(log_path,log_name))
    def write_log(summary_writer, logs_dict, epoch):
        with summary_writer.as_default():
            for key, value in logs_dict.items():
                print(key, value)
                tf.summary.scalar(key, data=value, step=epoch)

    
    def lr_decay(model, epoch, method,warm_epoch=0,max_epoch=40, bg_epoch=0, step_each_epoch=540, initial_lr=0.0001, power=0.9):
        ite = K.get_value(model.optimizer.iterations) + bg_epoch*step_each_epoch
        if ite < warm_epoch*step_each_epoch:
            l = ite*initial_lr/(step_each_epoch*warm_epoch)
        else:
            l = K.get_value(model.optimizer.lr) 
        if method == 'step_decay':  #step_decay
            if ite == round(max_epoch * 0.5)*step_each_epoch:
                l = l * 0.3
            elif ite == round(max_epoch * 0.7)*step_each_epoch:
                l = l * 0.3
            elif ite == round(max_epoch * 0.9)*step_each_epoch:
                l = l * 0.3
        elif method == 'poly':            
            l =  initial_lr*((1 - (ite / float(max_epoch*step_each_epoch)))**power)
            if l <= 1e-7:
                l = 1e-7
        return l
    def weight_decay(model, epoch, method,warm_epoch=0, max_epoch=40, bg_epoch=0, step_each_epoch=540, initial_wd=0.0001, power=0.9):
        ite = K.get_value(model.optimizer.iterations) + bg_epoch*step_each_epoch 
        if ite < warm_epoch*step_each_epoch:
            wd = ite*initial_wd/(step_each_epoch*warm_epoch)
        else:
            wd = K.get_value(model.optimizer.weight_decay) 
        if method == 'step_decay':  #step_decay
            if ite == round(max_epoch * 0.5)*step_each_epoch:
                wd = wd * 0.3
            elif ite == round(max_epoch * 0.7)*step_each_epoch:
                wd = wd * 0.3
            elif ite == round(max_epoch * 0.9)*step_each_epoch:
                wd = wd * 0.3
        elif method == 'poly': 
            ite = K.get_value(model.optimizer.weight_decay)       
            wd =  initial_wd*((1 - (ite / float(max_epoch*step_each_epoch)))**power)
            if wd <= 1e-7:
                wd = 1e-7
        return wd
        # IDRID:
    """
       Training:

       """
    class_weight = None
    log_name = log_name
    if not os.path.exists(log_path):
        os.makedirs(log_path)
    model_checkpoint = ModelCheckpoint(save_path + '/dp3_lesion.hdf5',
                                       monitor='val_mf1',
                                       verbose=1,
                                       mode='max',
                                       save_best_only=True,
                                       save_weights_only=True)
    # load weights

    if step != 2:    #loading initial weights(step=1),without initial weights(step=0)
        
        model.compile(loss=loss,
                          # loss_weights = loss_weights,
                          optimizer=opt,
                          # metrics = ['accuracy',IoU,IoUEX,IoUHE])   #注意修改
                          # metrics=['accuracy',IoU,realIoU,IoUEX,IoUHE,IoUMA,IoUSE,P_EX,R_EX,P_HE,R_HE,P_MA,R_MA,P_SE,R_SE])
                          metrics=['accuracy', IoU, mf1, IoUEX,
                                   IoUMA,f1_EX,
                                   f1_MA, P_EX,R_EX,P_MA,R_MA])
        
        if step == 1: 

            model.load_weights(load_dir_initial, by_name=True)
            print('model loaded!')
            model.save_weights(os.path.join(save_path, 'ep_lesion_initial.hdf5'))
            print('model saved!')
        elif step == 0:
            model.save_weights(os.path.join(save_path, 'ep_lesion_initial.hdf5'))
            print('model saved!')



        callback = TensorBoard(os.path.join(log_path, log_name))
        callback.set_model(model)


        train_names = ['train_loss', 'train_acc', 'train_mIoU',
                       'train_mf1', 'train_IoUEX',
                       'train_IoUMA',  'train_f1_EX',
                        'train_f1_MA',
                       'train_P_EX','train_R_EX',
                       'train_P_MA','train_R_MA','lr']
        val_names = ['val_loss', 'val_acc', 'val_mIoU',
                     'val_mf1', 'val_IoUEX', 
                     'val_IoUMA', 'val_f1_EX',
                      'val_f1_MA', 
                     'val_P_EX','val_R_EX','val_P_MA','val_R_MA',
                     'EX_AUPR',  'MA_AUPR', 'mAUPR']
        x1 = trainGene
        y1 = valGene
        loss_list = []
        PR_area = []
        MA_PR_area = []
        best_ep = 0
        alpha = [0.5,0.5,0.5]
        beta = [0.5,0.5,0.5]
        for i in range(bg_epoch, epochs, 1):
            train_loss_on_epoch = {}
            val_loss_on_epoch = {}
            for k1 in train_names:
                train_loss_on_epoch[k1] = 0  # create log-dict on epochs for training
            for k2 in val_names:
                val_loss_on_epoch[k2] = 0  # create log-dict on epochs for evaluation
            #current_lr = lr_decay(model, i, train_strategy, max_epoch = epochs, initial_lr=learning_rate, step_each_epoch=train_num)
            #K.set_value(model.optimizer.lr, current_lr)
            #current_wd = weight_decay(model, i, train_strategy, max_epoch = epochs, initial_wd=weight_decay_rate, step_each_epoch=train_num)
            #K.set_value(model.optimizer.weight_decay, current_wd)
            for j in range(math.ceil(train_num / train_batch_size)):
                l = K.get_value(model.optimizer.lr)
                current_lr = lr_decay(model, i, train_strategy, max_epoch = epochs, bg_epoch=bg_epoch, initial_lr=learning_rate, step_each_epoch=math.ceil(train_num / train_batch_size))
                K.set_value(model.optimizer.lr, current_lr)
                current_wd = weight_decay(model, i, train_strategy, max_epoch = epochs, bg_epoch=bg_epoch, initial_wd=weight_decay_rate, step_each_epoch=math.ceil(train_num / train_batch_size))
                K.set_value(model.optimizer.weight_decay, current_wd)
                train_x_y = next(x1)
                train_x, train_y = train_x_y[0], train_x_y[1]
                loss_on_batch = model.train_on_batch(train_x, train_y,return_dict=False)
                for m in range(len(loss_on_batch)):
                    train_loss_on_epoch[train_names[m]] = train_loss_on_epoch[train_names[m]] + loss_on_batch[m]
                print('ep:%d,%d/%d,lr:%f,loss:%f'%(i,j,math.ceil(train_num / train_batch_size),l,loss_on_batch[0]))
            for m in range(len(loss_on_batch)):  # calculate mean logs on per epoch
                train_loss_on_epoch[train_names[m]] = train_loss_on_epoch[train_names[m]] / train_num * train_batch_size
            print('ep:%d/%d'%(i, epochs), '\n', train_loss_on_epoch)
            #write_log(callback, train_loss_on_epoch, i)
            train_loss_on_epoch['lr'] = K.get_value(model.optimizer.lr)  #记录leraning rate
            write_log(summary_writer, train_loss_on_epoch, i)
            for n in range(math.ceil(val_num / val_batch_size)):
                val_x_y = next(y1)
                val_x, val_y = val_x_y[0], val_x_y[1]
                #y_pred = tf.compat.v1.to_double(model.predict_on_batch(val_x))
                #val_R = recall(val_y, y_pred)   
                #val_P = precision(val_y, y_pred)
                #val_R_EX,val_P_EX,val_R_HE,val_P_HE,val_R_MA,val_P_MA,val_R_SE,val_P_SE = val_R[1],val_P[1],val_R[2],val_P[2],val_R[3],val_P[3],val_R[4],val_P[4]
                val_loss_on_batch = model.test_on_batch(val_x, val_y,return_dict=False)
                real_val_num=np.ones(len(val_loss_on_batch))*val_num 
                for m in range(len(loss_on_batch)):
                    val_loss_on_epoch[val_names[m]] = val_loss_on_epoch[val_names[m]] + val_loss_on_batch[m]
            for m in range(len(loss_on_batch)):  # calculate mean logs on per epoch
                val_loss_on_epoch[val_names[m]] = val_loss_on_epoch[val_names[m]] / real_val_num[m] * val_batch_size
            loss_list.append(val_loss_on_epoch['val_loss'])
            loss_list_tmp = sorted(loss_list, reverse=False)
            #计算recall和precession，更新α和β
            alpha[1] = val_loss_on_epoch['val_R_EX']/(val_loss_on_epoch['val_R_EX']+val_loss_on_epoch['val_P_EX']+1e-6)
            alpha[2] = val_loss_on_epoch['val_R_MA']/(val_loss_on_epoch['val_R_MA']+val_loss_on_epoch['val_P_MA']+1e-6)
            for l in range(len(alpha)):   
                alpha[l] = 0.5+0.5*math.sin(3.14*(alpha[l]-0.5))        
                if alpha[l] < 0.1:
                    alpha[l] = 0.1
                elif alpha[l] > 0.9:
                    alpha[l] = 0.9
                beta[l] = 1 - alpha[l]
            #testGene = testGenerator(test_path, target_size=target_size)  # 读取验证集图片
            #results = model.predict_generator(testGene, val_num, verbose=1)
            '''
            if step == 0 and i == 0:
                model.save_weights(os.path.join(save_path, 'ep_lesion_initial.hdf5'))
                print('model saved！')
            '''
            
            if val_metrics:
                results = np.empty((val_num,target_size[0],target_size[1],classes),np.float32)
                testGene = testGenerator(test_path, target_size=target_size)  # 读取验证集图片
                for n in range(math.ceil(val_num / val_batch_size)):
                    val_x = next(testGene)
                    y_pred = model.predict_on_batch(val_x)
                    y_pred = tf.compat.v1.image.resize(y_pred, (target_size[0], target_size[1]),method=tf.image.ResizeMethod.NEAREST_NEIGHBOR)
                    print('val', y_pred.shape, y_pred.dtype)
                    results[n] = np.squeeze(y_pred)

            #val_loss_dict[val_names[len(loss_on_batch)] + ':'] = val_loss_on_epoch[len(loss_on_batch)]
            print('ep:%d/%d'%(i, epochs),'\n', val_loss_on_epoch)
            #write_log(callback, val_loss_on_epoch, i)
            write_log(summary_writer, val_loss_on_epoch, i)

        model.save_weights(os.path.join(save_path, 'ep_lesion_final.hdf5'))

        #cuda.select_device(int(GPU_id)) # 选择GPU设备
        #cuda.close() # 释放GPU资源

    else:  # 载入训练好的权重 step=2
        model.load_weights(load_dir, by_name=True)
        print('model loaded!')

    """
    test and visualize results
    
    """
    testGene = testGenerator(test_path, target_size=target_size)  # 读取验证集图片
    model.load_weights(load_dir, by_name=True)
    #best_results = model.predict_generator(testGene, val_num, verbose=1)
    z1 = testGene
    #best_results = model.predict_generator(testGene, val_num, verbose=1)
    best_results = np.empty((val_num,target_size[0],target_size[1],classes),np.float32)
    epsilon = np.empty(math.ceil(val_num / val_batch_size))
    for n in range(math.ceil(val_num / val_batch_size)):
        test_x_y = next(z1)
        test_x = test_x_y
        start = time.time()
        y_pred = model.predict_on_batch(test_x)
        epsilon[n] = (time.time() - start)
        print('Time used:', np.mean(epsilon[1:]))
        best_results[n] = np.squeeze(y_pred)
    if save_result == True:
        saveResult(save_path,
                   test_path,
                   target_size,
                   best_results,
                   flag_multi_class = flag_multi_class,
                   classes = classes,
                   task = task)  # 保存预测结果

        drawmask(test_path, save_path, classes = classes,target_size=target_size,task=task)  # visualization
        drawmask_truth(test_path, groundtruth_path, save_path, classes = classes,target_size=target_size,task=task)


    au_dict = evl(best_results,
                method = 'all',
                threshold_num = 2000,
                classes = classes,
                target_size = target_size,
                task = task,
                flag_multi_class = True,
                groundtruth_path = groundtruth_path,
                save_path = save_path)
    print(au_dict)
    




def main(args):

    #selecting a gpu
    GPU_id = args.GPU_id
    os.environ["CUDA_VISIBLE_DEVICES"] = GPU_id
    # tf.compat.v1.disable_eager_execution()
    config = tf.compat.v1.ConfigProto()
    config.gpu_options.allow_growth = True
    sess = tf.compat.v1.Session(config=config)

    step = args.step  # 0 for training
    use_pb = True if args.use_pb == 'y' else False
    classes = args.classes
    target_size = (args.target_size, args.target_size)
    train_batch_size = args.train_batch_size
    val_batch_size = args.val_batch_size  # keras只显示验证集每个batch的loss而非整个epoch的平均值，因此在不修改后端代码的情况下应设置较大的val_batchsize
    val_num = args.val_num
    train_num = args.train_num
    extra_aug = True if args.extra_aug == 'y' else False
    flag_multi_class = True if args.flag_multi_class == 'y' else False
    epochs = args.epochs  # when bs=1, epochs = 28
    bg_epoch = 0
    learning_rate = args.learning_rate
    weight_decay_rate = args.weight_decay_rate
    model_name = args.model_name  # unet_plusplus,unet3_plus,original_unet_plusplus,deeplabv3, unet
    supervision = True if args.supervision == 'y' else False
    use_ps = True if args.use_ps == 'y' else False
    normalize = args.norm  # BN, GN
    loss_f = args.loss
    train_strategy = args.train_strategy
    save_result = True if args.save_result == 'y' else False
    task = args.task
    if args.log_name:
        log_name = args.log_name
    else:
        log_name = args.dataset + '_' + str(args.target_size) + '_bs_' + str(
            train_batch_size) + '_reswunet++_' + normalize + '_adamw_2classes_pb_aug20_EX60_HE0_MA100_SE0_test1'
    log_path = args.dataset + '/tmp/log_pb_dense/'
    train_path = args.dataset + '/train/2 classes'
    val_path = args.dataset + '/val/2 classes'
    test_path = args.dataset + '/val/2 classes/image_zoom_hd'
    groundtruth_path = args.dataset + '/val/label_zoom_hd'
    save_path = args.dataset + '/result/2 classes/' + log_name
    load_dir_initial = args.load_dir_initial
    folder_dict = {
        0: [args.image_folder, args.label_folder],
        1: ['image_zoom_hd', 'label_zoom_hd']
    }

    trainGene, valGene = data_generator(classes = classes,
                                        use_pb=use_pb,
                                        folder_dict=folder_dict,
                                        target_size = target_size,
                                        train_batch_size = train_batch_size,
                                        val_batch_size = val_batch_size,       #keras只显示验证集每个batch的loss而非整个epoch的平均值，因此在不修改后端代码的情况下应设置较大的val_batchsize
                                        flag_multi_class = flag_multi_class,
                                        train_path = train_path,
                                        val_path = val_path,
                                        save_path = save_path,
                                        extra_aug = extra_aug)

    model, opt, loss = build_model(classes = classes,
                                    target_size = target_size,
                                    epochs = epochs,
                                    learning_rate = learning_rate,
                                    weight_decay_rate = weight_decay_rate,
                                    model_name = model_name,
                                    supervision = supervision,
                                    use_ps = use_ps,
                                    normalize = normalize,
                                    loss_f = loss_f)

    train(
          model, opt, loss,
          trainGene,
          valGene,
          GPU_id = GPU_id,
          train_strategy = train_strategy,
          step = step,
          classes = classes,
          target_size = target_size,
          train_batch_size = train_batch_size,
          val_batch_size = val_batch_size,       #keras只显示验证集每个batch的loss而非整个epoch的平均值，因此在不修改后端代码的情况下应设置较大的val_batchsize
          val_num = val_num,
          train_num = train_num,
          epochs = epochs,
          bg_epoch = bg_epoch,
          learning_rate = learning_rate,
          weight_decay_rate = weight_decay_rate,
          save_result = save_result,
          task = task,
          flag_multi_class = flag_multi_class,
          log_name = log_name,
          log_path= log_path,
          test_path = test_path,
          groundtruth_path = groundtruth_path,
          save_path = save_path,
          load_dir_initial = load_dir_initial
          )

import argparse
if __name__ == '__main__':
    paraser = argparse.ArgumentParser()
    paraser.add_argument('--GPU_id', type=str, default='0', help='GPU id')
    paraser.add_argument('--step', type=int, default=2, help='0 denotes training with random initial weight,'
                                                             '1 denotes training with a fixed initial weight, '
                                                             '2 denotes testing only')
    paraser.add_argument('--log_name', type=str,
                         default='e_ophtha_1024_bs_1_reswunet++_gn_adamw_2classes_pb_aug20_EX60_MA100_test1',
                         help='a log name for an exp')
    paraser.add_argument('--use_pb', type=str, default='yes', help='wo(yes) or w/o(no) PBDA')
    paraser.add_argument('--classes', type=int, default=3, help='number of label class')
    paraser.add_argument('--target_size', type=int, default=1024, help='input size of image')
    paraser.add_argument('--train_batch_size', type=int, default=1, help='batch size for training')
    paraser.add_argument('--val_batch_size', type=int, default=1, help='batch size for test')
    paraser.add_argument('--val_num', type=int, default=7, help='iterations of testing in one epoch')
    paraser.add_argument('--train_num', type=int, default=140, help='iterations of training in one epoch')
    paraser.add_argument('--extra_aug', type=str, default='y', help='extra augmentation methods')
    paraser.add_argument('--flag_multi_class', type=str, default='y', help='classes > 1?')
    paraser.add_argument('--epochs', type=int, default=28, help='maximum epochs for training')
    paraser.add_argument('--bg_epoch', type=int, default=0, help='beginning epoch number')
    paraser.add_argument('--learning_rate', type=int, default=0.0001, help='learning rate')
    paraser.add_argument('--weight_decay_rate', type=int, default=0.0001, help='weight decay rate')
    paraser.add_argument('--model_name', type=str, default='unet_plusplus',
                         help='model name: unet_plusplus/unet/deeplabv3+/unet3_plus/original_unet_plusplus')
    paraser.add_argument('--supervision', type=str, default='y', help='using supervision')
    paraser.add_argument('--use_ps', type=str, default='y', help='using pixel-shuffling')
    paraser.add_argument('--norm', type=str, default='gn', help='bn/gn')
    paraser.add_argument('--loss', type=str, default='dice_CE', help='loss function')
    paraser.add_argument('--train_strategy', type=str, default='step_decay', help='lr decay strategy')
    paraser.add_argument('--save_result', type=str, default='y', help='save image results?')
    paraser.add_argument('--task', type=str, default='ex_ma', help='task')
    paraser.add_argument('--load_dir_initial', type=str,
                         default='e_ophtha/result/2 classes/e_ophtha_1024_bs_1_reswunet++_gn_adamw_2classes_pb_aug20_EX60_MA100_test1/ep_lesion_initial.hdf5',
                         help='path of fixed initial weight')
    paraser.add_argument('--label_folder', type=str,
                         default='label_zoom_blend_hd_aug20_EX60_HE0_MA100_SE0',
                         help='name of label folder')
    paraser.add_argument('--image_folder', type=str,
                         default='image_zoom_blend_hd_aug20_EX60_HE0_MA100_SE0',
                         help='name of image folder')
    args = paraser.parse_args()
    #e_ophtha:
    main(args)


# In[ ]:




